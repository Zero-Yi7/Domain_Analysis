import xlwt
import socket
import openpyxl

wb_keywords = openpyxl.load_workbook("防护关键词-V2.0.xlsx")
ws_keywords = wb_keywords.active
keywords = [
    (row[0].value, row[1].value)
    for row in ws_keywords.iter_rows(min_row=2, max_col=2)
    if row[0].value
]

workbook = xlwt.Workbook()
worksheet = workbook.add_sheet("CNAME")

worksheet.write(0, 0, "Domain")
worksheet.write(0, 1, "Cname")
worksheet.write(0, 2, "服务分析")
worksheet.write(0, 3, "Aliases")

with open("domains.txt", "r", encoding="utf-8") as f:
    for i, line in enumerate(f):
        domain = line.strip()

        try:
            result = socket.gethostbyname_ex(domain)
            name = result[0]
            aliases = result[1]

            worksheet.write(i + 1, 0, domain)
            worksheet.write(i + 1, 1, name)

            if not domain.replace(".", "").isdigit():
                # 判断是否是IP地址
                is_cdn = any(keyword[0] in name.lower() for keyword in keywords)

                if is_cdn:
                    for row_index, keyword in enumerate(keywords, start=1):
                        if keyword[0] in name.lower():
                            worksheet.write(i + 1, 2, keyword[1])
                            break
                        elif row_index == len(keywords):
                            worksheet.write(i + 1, 2, "否")
                else:
                    worksheet.write(i + 1, 2, "否")
            else:
                worksheet.write(i + 1, 2, "资产为IP,无从判断")

            # 写入 Aliases
            for k, alias in enumerate(aliases):
                worksheet.write(i + 1, k + 3, alias)

        except Exception as e:
            worksheet.write(i + 1, 0, domain)
            worksheet.write(i + 1, 1, "错误")
            worksheet.write(i + 1, 2, "Ping请求不到主机")
            worksheet.write(i + 1, 3, str(e))
workbook.save("是否存在防护.xlsx")
